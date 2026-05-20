"""Generate Backend Developer Handoff as Excel and PDF on the user's Desktop."""

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

DESKTOP = Path.home() / "OneDrive" / "Desktop"
if not DESKTOP.exists():
    DESKTOP = Path.home() / "Desktop"

OUTPUT_XLSX = DESKTOP / "InsureEase_Backend_Developer_Handoff.xlsx"
OUTPUT_PDF = DESKTOP / "InsureEase_Backend_Developer_Handoff.pdf"

HEADER_FILL = PatternFill("solid", fgColor="1E40AF")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1E3A8A")
WRAP = Alignment(wrap_text=True, vertical="top")


def style_header_row(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def auto_width(ws, min_width=12, max_width=60):
    for col in ws.columns:
        letter = col[0].column_letter
        length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[letter].width = min(max(length + 2, min_width), max_width)


def add_sheet(ws, title, headers, rows):
    ws.title = title
    ws.append(headers)
    style_header_row(ws)
    for row in rows:
        ws.append(row)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP
    auto_width(ws)


def build_excel():
    wb = Workbook()

    # Overview
    ws = wb.active
    add_sheet(
        ws,
        "Overview",
        ["Section", "Details"],
        [
            ["Project", "InsureEase — Insurance comparison & lead capture platform"],
            ["Frontend", "React 19 + Vite 8 + React Router 7 (SPA)"],
            ["Live URL", "https://test-1-one-swart.vercel.app"],
            ["GitHub", "https://github.com/pavankumarbali/Insurance_company.git"],
            ["Current State", "Frontend-only prototype — no real API calls yet"],
            ["Validation", "src/utils/leadValidation.js"],
            ["Routes", "src/assets/App.jsx"],
            ["Generated", str(date.today())],
        ],
    )

    # API Endpoints
    ws2 = wb.create_sheet()
    add_sheet(
        ws2,
        "API Endpoints",
        ["Method", "Endpoint", "Purpose", "Priority"],
        [
            ["POST", "/api/v1/auth/create-account", "User signup", "High"],
            ["POST", "/api/v1/auth/send-otp", "Send login OTP to mobile", "High"],
            ["POST", "/api/v1/auth/verify-otp", "Verify OTP and create session", "High"],
            ["POST", "/api/v1/leads/health", "Save health insurance quote lead", "High"],
            ["POST", "/api/v1/leads/term", "Save term insurance quote lead", "High"],
            ["POST", "/api/v1/leads/motor", "Save motor insurance quote lead", "High"],
            ["POST", "/api/v1/leads/business", "Save business insurance lead", "High"],
            ["POST", "/api/v1/leads/cargo", "Save cargo insurance lead", "High"],
            ["POST", "/api/v1/leads/contact", "Save contact / callback request", "High"],
            ["GET", "/api/v1/vehicle/lookup?number=XX", "Vehicle registration lookup", "High"],
            ["POST", "/api/v1/vehicle/rc-extract", "RC book OCR / document extract", "Medium"],
            ["GET", "/api/v1/masters/brands-models", "Vehicle brand / model / variant master", "Medium"],
            ["GET", "/api/v1/plans?product=motor&category=car", "Insurance plan catalog", "Medium"],
            ["POST", "/api/v1/quotes/premium", "Premium calculation", "Medium"],
        ],
    )

    # Auth
    ws3 = wb.create_sheet()
    add_sheet(
        ws3,
        "Auth",
        ["Feature", "Current (Frontend)", "Backend Required", "Fields"],
        [
            ["Login OTP", "Demo OTP hardcoded as 1234", "Real OTP send + verify APIs", "mobileNumber, otp"],
            ["Signup", "Saved in React state only", "POST /api/v1/auth/create-account", "fullName, mobileNumber, email, password"],
            ["Guest login", "Navigates home, no session", "Optional anonymous session", "—"],
            ["Sessions", "None", "JWT or HTTP-only cookies", "token, userId, expiresAt"],
        ],
    )

    # Forms
    ws4 = wb.create_sheet()
    add_sheet(
        ws4,
        "Forms & Data",
        ["Product", "File", "Fields Collected", "Current Submit Behavior"],
        [
            ["Login", "Login/Login.jsx", "mobileNumber, otpCode", "Client-only; OTP = 1234"],
            ["Signup", "Signup/Signup.jsx", "fullName, mobileNumber, emailAddress, password", "Local state only"],
            ["Health", "HealthInsurance/Health-Home.jsx", "Members, contact, medical conditions", "Alert only"],
            ["Term", "Term/TermQuotePanel.jsx", "fullName, dateOfBirth, mobileNumber, gender, smoker", "In-memory leadQueue"],
            ["Motor (plate)", "MotorInsurance/MotorInsurance.jsx", "vehicleNumber, category", "Dummy policy lookup"],
            ["Motor (RC)", "MotorInsurance/MotorInsurance.jsx", "RC file upload", "Mock OCR after 900ms"],
            ["Motor (no number)", "Withoutnumber/WithoutNumber.jsx", "vehicle + insurance details", "console.info only"],
            ["Motor (new car)", "Newcar/Newcar.jsx", "brand, model, variant, fuel, city, name, mobile", "console log only"],
            ["Business", "BusinessInsurance/Business-Home.jsx", "fullName, mobileNumber, coverageType", "Navigate only"],
            ["Cargo", "cargo/Cargo-Merain|Air|Inland.jsx", "fullName, businessName, mobileNumber, email", "Alert only"],
            ["Contact", "Contact/ContactUs.jsx", "fullName, mobile, email, insuranceType, message", "No submit handler"],
        ],
    )

    # Routes
    ws5 = wb.create_sheet()
    add_sheet(
        ws5,
        "Routes",
        ["Path", "Page", "Component File"],
        [
            ["/", "Landing", "Landingpage/Landing-Page.jsx"],
            ["/login", "Login", "Login/Login.jsx"],
            ["/signup", "Signup", "Signup/Signup.jsx"],
            ["/motor-insurance/:category", "Motor", "MotorInsurance/MotorInsurance.jsx"],
            ["/health-insurance", "Health", "HealthInsurance/Health-Home.jsx"],
            ["/term-insurance", "Term", "Term/Term-Home.jsx"],
            ["/cargo-insurance", "Cargo hub", "cargo/Cargo-Home.jsx"],
            ["/cargo-insurance/marine|air|inland", "Cargo products", "cargo/Cargo-*.jsx"],
            ["/business-insurance", "Business hub", "BusinessInsurance/Business-Home.jsx"],
            ["/business/fire, /business-insurance/*", "Business details", "BusinessInsurance/*.jsx"],
            ["/contact-us", "Contact", "Contact/ContactUs.jsx"],
            ["/insurance-basics", "Education", "InsuranceBasics/InsuranceBasics.jsx"],
        ],
    )

    # Mock Data
    ws6 = wb.create_sheet()
    add_sheet(
        ws6,
        "Mock Data Files",
        ["File", "Purpose", "Replace With"],
        [
            ["MotorPolicyDummyData.jsx", "Fake policy from vehicle number", "Vehicle lookup API"],
            ["GuestDummyData.jsx", "Motor plans & add-ons", "Plan catalog API"],
            ["NewcarDummyData.jsx", "Brand/model/variant lists", "MMV master data API"],
            ["MotorInsurance.jsx buildDummyRcExtractedFields", "RC OCR simulation", "RC extract API"],
            ["Term-Home.jsx leadQueue", "Term leads in memory", "POST /api/v1/leads/term"],
            ["Signup.jsx createdAccounts", "Signup in memory", "POST /api/v1/auth/create-account"],
        ],
    )

    # Integration
    ws7 = wb.create_sheet()
    add_sheet(
        ws7,
        "Frontend Integration",
        ["Step", "Action"],
        [
            ["1", "Add VITE_API_BASE_URL environment variable"],
            ["2", "Create API service layer (fetch/axios)"],
            ["3", "Wire signup to POST /api/v1/auth/create-account"],
            ["4", "Replace OTP 1234 with real send/verify flow"],
            ["5", "POST form data on health, term, motor, business, cargo, contact submit"],
            ["6", "Replace dummy vehicle lookup with GET /api/v1/vehicle/lookup"],
            ["7", "Replace RC mock with POST /api/v1/vehicle/rc-extract"],
            ["8", "Add JWT/session token to authenticated requests"],
            ["9", "Enable CORS on backend for Vercel frontend domain"],
        ],
    )

    wb.save(OUTPUT_XLSX)
    print(f"Excel saved: {OUTPUT_XLSX}")


def build_pdf():
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "Title",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=18,
        textColor=colors.HexColor("#1E3A8A"),
        spaceAfter=12,
    )
    h2 = ParagraphStyle(
        "H2",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=13,
        textColor=colors.HexColor("#0F172A"),
        spaceBefore=10,
        spaceAfter=6,
    )
    body = ParagraphStyle(
        "Body",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=10,
        leading=14,
        spaceAfter=6,
    )

    story = [
        Paragraph("InsureEase — Backend Developer Handoff", title_style),
        Paragraph(f"Generated: {date.today()}", body),
        Spacer(1, 8),
        Paragraph("Project Overview", h2),
        Paragraph(
            "InsureEase is a React 19 + Vite 8 SPA for insurance comparison and lead capture. "
            "The frontend is complete as a UI prototype. No real API calls exist yet — all forms "
            "validate client-side and use dummy data, alerts, or console logs.",
            body,
        ),
        Paragraph("Live: https://test-1-one-swart.vercel.app", body),
        Paragraph("GitHub: https://github.com/pavankumarbali/Insurance_company.git", body),
        Paragraph("Priority APIs", h2),
    ]

    api_data = [
        ["Method", "Endpoint", "Purpose"],
        ["POST", "/api/v1/auth/create-account", "Signup"],
        ["POST", "/api/v1/auth/send-otp", "Send OTP"],
        ["POST", "/api/v1/auth/verify-otp", "Verify OTP + session"],
        ["POST", "/api/v1/leads/health", "Health quote lead"],
        ["POST", "/api/v1/leads/term", "Term quote lead"],
        ["POST", "/api/v1/leads/motor", "Motor quote lead"],
        ["POST", "/api/v1/leads/business", "Business lead"],
        ["POST", "/api/v1/leads/cargo", "Cargo lead"],
        ["POST", "/api/v1/leads/contact", "Contact form"],
        ["GET", "/api/v1/vehicle/lookup", "Vehicle registration lookup"],
        ["POST", "/api/v1/vehicle/rc-extract", "RC book OCR"],
        ["GET", "/api/v1/masters/brands-models", "Vehicle master data"],
        ["POST", "/api/v1/quotes/premium", "Premium calculator"],
    ]
    table = Table(api_data, colWidths=[50, 180, 200])
    table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E40AF")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ])
    )
    story.append(table)
    story.append(Spacer(1, 10))

    sections = [
        ("Authentication", [
            "Login uses hardcoded OTP 1234 — replace with real send/verify APIs.",
            "Signup fields: fullName, mobileNumber, email (optional), password (min 6).",
            "No JWT, cookies, or localStorage session today.",
        ]),
        ("Forms to Wire Up", [
            "Health: 3-step wizard — members, contact, medical conditions.",
            "Term: name, DOB, mobile, gender, smoker, WhatsApp opt-in.",
            "Motor: vehicle number, RC upload, without-number flow, new car wizard.",
            "Business: popup lead (name, mobile, coverage type).",
            "Cargo: marine/air/inland — name, business, mobile, email.",
            "Contact: name, mobile, email, insurance type, message.",
        ]),
        ("Key Frontend Files", [
            "Routes: src/assets/App.jsx",
            "Validation: src/utils/leadValidation.js",
            "Auth: src/assets/Login/Login.jsx, Signup/Signup.jsx",
            "Motor mocks: MotorPolicyDummyData.jsx, GuestDummyData.jsx, NewcarDummyData.jsx",
        ]),
        ("Integration Checklist", [
            "Add VITE_API_BASE_URL env variable.",
            "Create API service layer and wire all form submits.",
            "Replace dummy data with real API responses.",
            "Enable CORS for Vercel production domain.",
        ]),
    ]

    for heading, bullets in sections:
        story.append(Paragraph(heading, h2))
        for bullet in bullets:
            story.append(Paragraph(f"• {bullet}", body))

    doc = SimpleDocTemplate(
        str(OUTPUT_PDF),
        pagesize=A4,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=16 * mm,
        bottomMargin=16 * mm,
        title="InsureEase Backend Developer Handoff",
    )
    doc.build(story)
    print(f"PDF saved: {OUTPUT_PDF}")


def main():
    DESKTOP.mkdir(parents=True, exist_ok=True)
    build_excel()
    build_pdf()
    print("Done — files saved to Desktop.")


if __name__ == "__main__":
    main()
